"""Tests d'export Excel : fichier reellement ouvrable et contenu complet."""
import io

import pytest
from openpyxl import load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORTS = ["/api/export/candidats", "/api/export/postes", "/api/export/process"]


def _feuille(resp):
    """Ouvre la reponse comme un vrai classeur : valide le format de bout en bout."""
    return load_workbook(io.BytesIO(resp.content)).active


@pytest.mark.parametrize("chemin", EXPORTS)
def test_export_repond_un_xlsx(api, chemin):
    resp = api.get(chemin)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MIME


@pytest.mark.parametrize("chemin", EXPORTS)
def test_export_propose_un_telechargement(api, chemin):
    disposition = api.get(chemin).headers["content-disposition"]
    assert "attachment" in disposition
    assert ".xlsx" in disposition


@pytest.mark.parametrize("chemin", EXPORTS)
def test_export_vide_reste_ouvrable(api, chemin):
    """Base vide : le classeur doit exister avec seulement sa ligne d'en-tetes."""
    feuille = _feuille(api.get(chemin))
    assert feuille.max_row >= 1


@pytest.mark.parametrize("chemin", EXPORTS)
def test_export_sans_authentification(client, chemin):
    assert client.get(chemin).status_code in (401, 403)


def test_export_candidats_contient_toutes_les_lignes(api, make_candidat):
    for i in range(7):
        make_candidat(nom=f"Nom{i}", prenom=f"Pre{i}")

    feuille = _feuille(api.get("/api/export/candidats"))
    assert feuille.max_row == 8  # 7 candidats + en-tetes


def test_export_candidats_contient_les_valeurs(api, make_candidat):
    make_candidat(nom="Exportable", prenom="Alice", ville="Nantes")
    feuille = _feuille(api.get("/api/export/candidats"))

    valeurs = {str(c.value) for ligne in feuille.iter_rows() for c in ligne}
    assert "Exportable" in valeurs
    assert "Nantes" in valeurs


def test_export_candidats_inclut_les_archives(api, make_candidat):
    actif = make_candidat(nom="Actif")
    archive = make_candidat(nom="Archive")
    api.put(f"/api/candidats/{archive['id']}", json={"is_archived": True})

    feuille = _feuille(api.get("/api/export/candidats"))
    assert feuille.max_row == 3
    assert actif["id"] and archive["id"]


def test_export_postes_contient_toutes_les_lignes(api, make_poste):
    for i in range(5):
        make_poste(entreprise=f"Ent{i}")

    feuille = _feuille(api.get("/api/export/postes"))
    assert feuille.max_row == 6


def test_export_postes_contient_les_valeurs(api, make_poste):
    make_poste(entreprise="ExportCorp", ville="Bordeaux", contact="Sophie")
    valeurs = {str(c.value) for ligne in _feuille(api.get("/api/export/postes")).iter_rows()
               for c in ligne}
    assert "ExportCorp" in valeurs
    assert "Bordeaux" in valeurs


def test_export_process_contient_toutes_les_lignes(api, make_candidat, make_poste, make_process):
    poste = make_poste()
    for i in range(4):
        candidat = make_candidat(nom=f"Nom{i}")
        make_process(candidat["id"], poste["id"], statut="ENCV")

    feuille = _feuille(api.get("/api/export/process"))
    assert feuille.max_row == 5


def test_export_process_joint_candidat_et_poste(api, make_candidat, make_poste, make_process):
    candidat = make_candidat(nom="Leroy", prenom="Francois")
    poste = make_poste(entreprise="DataLab")
    make_process(candidat["id"], poste["id"], statut="PCLT", honoraire=7500.0)

    # Le nom du candidat est concatene dans une seule cellule : on cherche en sous-chaine
    texte = " | ".join(
        str(c.value) for ligne in _feuille(api.get("/api/export/process")).iter_rows()
        for c in ligne
    )
    assert "Leroy" in texte
    assert "DataLab" in texte


def test_export_au_dela_de_1000_lignes(api, sync_db):
    """Regression : .to_list(1000) tronquait aussi les exports."""
    import uuid
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc).isoformat()
    sync_db.candidats.insert_many([{
        "id": str(uuid.uuid4()), "nom": f"Nom{i:05d}", "prenom": "P", "ville": "Paris",
        "titre_poste": "Dev", "rayon_km": 30, "is_archived": False, "created_at": now,
    } for i in range(1200)])

    feuille = _feuille(api.get("/api/export/candidats"))
    assert feuille.max_row == 1201
