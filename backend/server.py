from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import math
import io
import httpx

# Cache pour les coordonnées des villes (évite de rappeler l'API)
city_coords_cache = {}

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'recruithub-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ============ MODELS ============

class UserCreate(BaseModel):
    email: str
    password: str

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# Statuts possibles pour un candidat
STATUTS_CANDIDAT = ["NOUVEAU", "ENCV", "ENTC", "PROPALE", "PCLT", "REFUS", "NOGO_DISPO"]
SOURCES_CANDIDAT = [
    "Hellowork candidature",
    "Hellowork cvtech", 
    "Indeed",
    "LinkedIn",
    "Site 111 conseils",
    "Cooptation"
]

# Candidat Models
class CandidatCreate(BaseModel):
    nom: str
    prenom: str
    ville: str
    code_postal: Optional[str] = None
    rayon_km: int = 30
    titre_poste: str
    remuneration: Optional[str] = None
    disponibilite: Optional[str] = None
    source: Optional[str] = None

class CandidatUpdate(BaseModel):
    nom: Optional[str] = None
    prenom: Optional[str] = None
    ville: Optional[str] = None
    code_postal: Optional[str] = None
    rayon_km: Optional[int] = None
    titre_poste: Optional[str] = None
    remuneration: Optional[str] = None
    disponibilite: Optional[str] = None
    source: Optional[str] = None

class Candidat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    prenom: str
    ville: str
    code_postal: Optional[str] = None
    rayon_km: int = 30
    titre_poste: str
    remuneration: Optional[str] = None
    disponibilite: Optional[str] = None
    source: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Process Models (suivi candidat-poste)
class ProcessCreate(BaseModel):
    candidat_id: str
    poste_id: str
    statut: str = "ENCV"
    honoraire: Optional[float] = None
    notes: Optional[str] = None

class ProcessUpdate(BaseModel):
    statut: Optional[str] = None
    honoraire: Optional[float] = None
    notes: Optional[str] = None

class Process(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    candidat_id: str
    poste_id: str
    statut: str = "ENCV"
    honoraire: Optional[float] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Poste Models
class PosteCreate(BaseModel):
    entreprise: str
    titre_poste: str
    ville: str
    code_postal: Optional[str] = None
    convention_signee: bool = False
    contact: Optional[str] = None
    email_contact: Optional[str] = None

class PosteUpdate(BaseModel):
    entreprise: Optional[str] = None
    titre_poste: Optional[str] = None
    ville: Optional[str] = None
    code_postal: Optional[str] = None
    convention_signee: Optional[bool] = None
    contact: Optional[str] = None
    email_contact: Optional[str] = None

class Poste(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entreprise: str
    titre_poste: str
    ville: str
    code_postal: Optional[str] = None
    convention_signee: bool = False
    contact: Optional[str] = None
    email_contact: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Match Model
class Match(BaseModel):
    candidat: Candidat
    poste_id: str
    score: int
    titre_match: bool
    zone_match: bool

# ============ FRENCH CITIES COORDINATES ============

FRENCH_CITIES = {
    # Grandes villes
    "paris": (48.8566, 2.3522),
    "marseille": (43.2965, 5.3698),
    "lyon": (45.7640, 4.8357),
    "toulouse": (43.6047, 1.4442),
    "nice": (43.7102, 7.2620),
    "nantes": (47.2184, -1.5536),
    "strasbourg": (48.5734, 7.7521),
    "montpellier": (43.6108, 3.8767),
    "bordeaux": (44.8378, -0.5792),
    "lille": (50.6292, 3.0573),
    "rennes": (48.1173, -1.6778),
    "reims": (49.2583, 4.0317),
    "saint-etienne": (45.4397, 4.3872),
    "toulon": (43.1242, 5.9280),
    "le havre": (49.4944, 0.1079),
    "grenoble": (45.1885, 5.7245),
    "dijon": (47.3220, 5.0415),
    "angers": (47.4784, -0.5632),
    "nimes": (43.8367, 4.3601),
    "villeurbanne": (45.7676, 4.8798),
    "clermont-ferrand": (45.7772, 3.0870),
    "le mans": (48.0061, 0.1996),
    "aix-en-provence": (43.5297, 5.4474),
    "brest": (48.3904, -4.4861),
    "tours": (47.3941, 0.6848),
    "amiens": (49.8941, 2.2958),
    "limoges": (45.8336, 1.2611),
    "perpignan": (42.6986, 2.8954),
    "besancon": (47.2378, 6.0241),
    "orleans": (47.9029, 1.9093),
    "metz": (49.1193, 6.1757),
    "rouen": (49.4432, 1.0993),
    "mulhouse": (47.7508, 7.3359),
    "caen": (49.1829, -0.3707),
    "nancy": (48.6921, 6.1844),
    "argenteuil": (48.9472, 2.2467),
    "saint-denis": (48.9362, 2.3574),
    "roubaix": (50.6942, 3.1746),
    "tourcoing": (50.7262, 3.1612),
    "montreuil": (48.8638, 2.4483),
    "avignon": (43.9493, 4.8055),
    "dunkerque": (51.0343, 2.3768),
    "poitiers": (46.5802, 0.3404),
    "versailles": (48.8014, 2.1301),
    "courbevoie": (48.8966, 2.2526),
    "creteil": (48.7909, 2.4551),
    "pau": (43.2951, -0.3708),
    "la rochelle": (46.1603, -1.1511),
    "cannes": (43.5528, 7.0174),
    "antibes": (43.5808, 7.1238),
    
    # Métropole Bordelaise
    "bruges": (44.8833, -0.6167),
    "merignac": (44.8386, -0.6436),
    "mérignac": (44.8386, -0.6436),
    "pessac": (44.8067, -0.6311),
    "talence": (44.8014, -0.5878),
    "begles": (44.8083, -0.5481),
    "bègles": (44.8083, -0.5481),
    "villenave-d'ornon": (44.7803, -0.5569),
    "gradignan": (44.7719, -0.6156),
    "le bouscat": (44.8656, -0.6019),
    "cenon": (44.8572, -0.5328),
    "lormont": (44.8761, -0.5206),
    "floirac": (44.8344, -0.5086),
    "blanquefort": (44.9128, -0.6381),
    "eysines": (44.8833, -0.6500),
    "saint-medard-en-jalles": (44.8958, -0.7178),
    "le haillan": (44.8667, -0.6833),
    "carbon-blanc": (44.8944, -0.5000),
    "ambares-et-lagrave": (44.9333, -0.4833),
    "bassens": (44.9000, -0.5167),
    "artigues-pres-bordeaux": (44.8500, -0.4833),
    "libourne": (44.9167, -0.2333),
    "arcachon": (44.6608, -1.1681),
    "la teste-de-buch": (44.6333, -1.1333),
    "gujan-mestras": (44.6333, -1.0667),
    
    # Métropole Lyonnaise
    "venissieux": (45.6972, 4.8864),
    "vénissieux": (45.6972, 4.8864),
    "saint-priest": (45.6969, 4.9428),
    "vaulx-en-velin": (45.7833, 4.9167),
    "bron": (45.7386, 4.9131),
    "caluire-et-cuire": (45.7958, 4.8464),
    "oullins": (45.7144, 4.8078),
    "tassin-la-demi-lune": (45.7639, 4.7794),
    "sainte-foy-les-lyon": (45.7333, 4.8000),
    "ecully": (45.7750, 4.7781),
    "rillieux-la-pape": (45.8167, 4.9000),
    "meyzieu": (45.7667, 5.0000),
    "decines-charpieu": (45.7667, 4.9500),
    
    # Métropole Lilloise
    "villeneuve-d'ascq": (50.6167, 3.1333),
    "wattrelos": (50.7000, 3.2167),
    "marcq-en-baroeul": (50.6667, 3.1000),
    "lambersart": (50.6500, 3.0333),
    "armentieres": (50.6833, 2.8833),
    "la madeleine": (50.6500, 3.0667),
    "mons-en-baroeul": (50.6333, 3.1000),
    "hem": (50.6500, 3.1833),
    "loos": (50.6167, 3.0167),
    "croix": (50.6667, 3.1500),
    
    # Métropole Toulousaine
    "colomiers": (43.6000, 1.3333),
    "tournefeuille": (43.5833, 1.3500),
    "blagnac": (43.6333, 1.3833),
    "muret": (43.4667, 1.3333),
    "ramonville-saint-agne": (43.5500, 1.4833),
    "cugnaux": (43.5333, 1.3500),
    "balma": (43.6000, 1.5000),
    "l'union": (43.6500, 1.4833),
    "saint-orens-de-gameville": (43.5500, 1.5333),
    "castanet-tolosan": (43.5167, 1.5000),
    
    # Métropole Nantaise
    "saint-herblain": (47.2167, -1.6500),
    "reze": (47.1833, -1.5500),
    "rezé": (47.1833, -1.5500),
    "saint-sebastien-sur-loire": (47.2000, -1.5000),
    "orvault": (47.2667, -1.6167),
    "vertou": (47.1667, -1.4667),
    "coueron": (47.2167, -1.7167),
    "carquefou": (47.2833, -1.4833),
    "la chapelle-sur-erdre": (47.3000, -1.5500),
    "bouguenais": (47.1667, -1.6167),
    
    # Ile-de-France
    "boulogne-billancourt": (48.8333, 2.2500),
    "saint-denis": (48.9362, 2.3574),
    "nanterre": (48.8922, 2.2069),
    "vitry-sur-seine": (48.7833, 2.4000),
    "aubervilliers": (48.9167, 2.3833),
    "colombes": (48.9167, 2.2500),
    "asnieres-sur-seine": (48.9167, 2.2833),
    "rueil-malmaison": (48.8833, 2.1833),
    "champigny-sur-marne": (48.8167, 2.5167),
    "saint-maur-des-fosses": (48.8000, 2.5000),
    "drancy": (48.9333, 2.4500),
    "issy-les-moulineaux": (48.8167, 2.2667),
    "levallois-perret": (48.8833, 2.2833),
    "noisy-le-grand": (48.8500, 2.5667),
    "antony": (48.7500, 2.3000),
    "neuilly-sur-seine": (48.8833, 2.2667),
    "clichy": (48.9000, 2.3000),
    "ivry-sur-seine": (48.8167, 2.3833),
    "pantin": (48.8833, 2.4000),
    "bondy": (48.9000, 2.4833),
    
    # Métropole Marseillaise
    "aix-en-provence": (43.5297, 5.4474),
    "martigues": (43.4000, 5.0500),
    "aubagne": (43.2833, 5.5667),
    "istres": (43.5167, 4.9833),
    "la ciotat": (43.1667, 5.6000),
    "vitrolles": (43.4500, 5.2500),
    "marignane": (43.4167, 5.2167),
    "salon-de-provence": (43.6333, 5.1000),
    
    # Autres
    "bayonne": (43.4833, -1.4833),
    "anglet": (43.4833, -1.5167),
    "biarritz": (43.4833, -1.5500),
    "saint-nazaire": (47.2833, -2.2000),
    "vannes": (47.6500, -2.7500),
    "lorient": (47.7500, -3.3667),
    "quimper": (48.0000, -4.1000),
    "saint-brieuc": (48.5167, -2.7667),
    "valence": (44.9333, 4.8833),
    "chambery": (45.5667, 5.9167),
    "annecy": (45.9000, 6.1167),
    "troyes": (48.3000, 4.0833),
    "bourges": (47.0833, 2.4000),
    "niort": (46.3167, -0.4667),
    "la roche-sur-yon": (46.6667, -1.4333),
    "chartres": (48.4500, 1.5000),
    "colmar": (48.0833, 7.3500),
    "chalons-en-champagne": (48.9500, 4.3667),
    "beauvais": (49.4333, 2.0833),
    "saint-quentin": (49.8500, 3.2833),
    "laval": (48.0667, -0.7667),
    "belfort": (47.6333, 6.8667),
    "tarbes": (43.2333, 0.0833),
    "albi": (43.9333, 2.1500),
    "montauban": (44.0167, 1.3500),
    "beziers": (43.3500, 3.2167),
    "narbonne": (43.1833, 3.0000),
    "carcassonne": (43.2167, 2.3500),
    "rodez": (44.3500, 2.5667),
    "cahors": (44.4500, 1.4333),
    "agen": (44.2000, 0.6167),
    "perigueux": (45.1833, 0.7167),
    "bergerac": (44.8500, 0.4833),
    "mont-de-marsan": (43.8833, -0.5000),
    "dax": (43.7167, -1.0500),
    "angouleme": (45.6500, 0.1500),
    "cognac": (45.7000, -0.3333),
    "rochefort": (45.9333, -0.9667),
    "saintes": (45.7500, -0.6333),
    "royan": (45.6167, -1.0333),
}

async def get_city_coords_from_api(city_name: str, code_postal: str = None) -> Optional[tuple]:
    """Récupère les coordonnées d'une ville via l'API Adresse du gouvernement"""
    try:
        query = f"{city_name} {code_postal}" if code_postal else city_name
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(
                "https://api-adresse.data.gouv.fr/search/",
                params={"q": query, "type": "municipality", "limit": 1}
            )
            if response.status_code == 200:
                data = response.json()
                if data.get("features") and len(data["features"]) > 0:
                    coords = data["features"][0]["geometry"]["coordinates"]
                    # L'API retourne [longitude, latitude], on inverse
                    return (coords[1], coords[0])
    except Exception as e:
        logging.warning(f"Erreur géocodage {city_name}: {e}")
    return None

def get_city_coords(city_name: str) -> Optional[tuple]:
    """Get coordinates for a city name - version synchrone pour compatibilité"""
    normalized = city_name.lower().strip()
    
    # Vérifier le cache d'abord
    if normalized in city_coords_cache:
        return city_coords_cache[normalized]
    
    # Vérifier dans la liste locale
    if normalized in FRENCH_CITIES:
        return FRENCH_CITIES[normalized]
    
    # Try partial match dans la liste locale
    for city, coords in FRENCH_CITIES.items():
        if normalized in city or city in normalized:
            return coords
    
    return None

async def get_city_coords_async(city_name: str, code_postal: str = None) -> Optional[tuple]:
    """Get coordinates for a city name - version asynchrone avec API"""
    cache_key = f"{city_name.lower().strip()}_{code_postal or ''}"
    
    # Vérifier le cache d'abord
    if cache_key in city_coords_cache:
        return city_coords_cache[cache_key]
    
    normalized = city_name.lower().strip()
    
    # Vérifier dans la liste locale
    if normalized in FRENCH_CITIES:
        city_coords_cache[cache_key] = FRENCH_CITIES[normalized]
        return FRENCH_CITIES[normalized]
    
    # Try partial match dans la liste locale
    for city, coords in FRENCH_CITIES.items():
        if normalized in city or city in normalized:
            city_coords_cache[cache_key] = coords
            return coords
    
    # Appeler l'API pour les villes non trouvées
    coords = await get_city_coords_from_api(city_name, code_postal)
    if coords:
        city_coords_cache[cache_key] = coords
        return coords
    
    return None

def calculate_distance_km(coord1: tuple, coord2: tuple) -> float:
    """Calculate distance between two coordinates using Haversine formula"""
    lat1, lon1 = coord1
    lat2, lon2 = coord2
    R = 6371  # Earth's radius in km
    
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    return R * c

def normalize_title(title: str) -> str:
    """Normalize job title for comparison"""
    return title.lower().strip()

def titles_match(title1: str, title2: str) -> bool:
    """Check if two job titles match (fuzzy)"""
    t1 = normalize_title(title1)
    t2 = normalize_title(title2)
    # Exact match or one contains the other
    return t1 == t2 or t1 in t2 or t2 in t1

def calculate_match_score(candidat: dict, poste: dict) -> dict:
    """Calculate matching score between a candidate and a job position - version synchrone"""
    score = 0
    titre_match = False
    zone_match = False
    
    # Title matching (50% weight)
    if titles_match(candidat['titre_poste'], poste['titre_poste']):
        score += 50
        titre_match = True
    
    # Geographic matching (50% weight)
    candidat_coords = get_city_coords(candidat['ville'])
    poste_coords = get_city_coords(poste['ville'])
    
    if candidat_coords and poste_coords:
        distance = calculate_distance_km(candidat_coords, poste_coords)
        rayon = candidat['rayon_km']
        
        if distance <= rayon:
            # Full score if within radius
            zone_match = True
            # Score decreases linearly with distance
            zone_score = max(0, 50 * (1 - distance / rayon))
            score += zone_score
    elif candidat['ville'].lower().strip() == poste['ville'].lower().strip():
        # Same city name
        score += 50
        zone_match = True
    
    return {
        'score': int(score),
        'titre_match': titre_match,
        'zone_match': zone_match
    }

async def calculate_match_score_async(candidat: dict, poste: dict) -> dict:
    """Calculate matching score - version asynchrone avec géocodage API"""
    score = 0
    titre_match = False
    zone_match = False
    
    # Title matching (50% weight)
    if titles_match(candidat['titre_poste'], poste['titre_poste']):
        score += 50
        titre_match = True
    
    # Geographic matching (50% weight)
    candidat_coords = await get_city_coords_async(candidat['ville'], candidat.get('code_postal'))
    poste_coords = await get_city_coords_async(poste['ville'], poste.get('code_postal'))
    
    if candidat_coords and poste_coords:
        distance = calculate_distance_km(candidat_coords, poste_coords)
        rayon = candidat.get('rayon_km', 30)
        
        if distance <= rayon:
            zone_match = True
            zone_score = max(0, 50 * (1 - distance / rayon))
            score += zone_score
    elif candidat['ville'].lower().strip() == poste['ville'].lower().strip():
        score += 50
        zone_match = True
    
    return {
        'score': int(score),
        'titre_match': titre_match,
        'zone_match': zone_match
    }

# ============ AUTH HELPERS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('user_id')
        if not user_id:
            raise HTTPException(status_code=401, detail="Token invalide")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    # Inscription désactivée - application privée
    raise HTTPException(status_code=403, detail="Les inscriptions sont fermées. Contactez l'administrateur.")

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user: UserLogin):
    user_doc = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not user_doc or not verify_password(user.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    token = create_token(user_doc['id'], user_doc['email'])
    return TokenResponse(
        access_token=token,
        user=UserResponse(id=user_doc['id'], email=user_doc['email'])
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(id=current_user['user_id'], email=current_user['email'])

# ============ CANDIDATS ROUTES ============

@api_router.get("/candidats", response_model=List[Candidat])
async def get_candidats(current_user: dict = Depends(get_current_user)):
    candidats = await db.candidats.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for c in candidats:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return candidats

@api_router.get("/candidats/{candidat_id}", response_model=Candidat)
async def get_candidat(candidat_id: str, current_user: dict = Depends(get_current_user)):
    candidat = await db.candidats.find_one({"id": candidat_id}, {"_id": 0})
    if not candidat:
        raise HTTPException(status_code=404, detail="Candidat non trouvé")
    if isinstance(candidat.get('created_at'), str):
        candidat['created_at'] = datetime.fromisoformat(candidat['created_at'])
    return candidat

@api_router.post("/candidats", response_model=Candidat)
async def create_candidat(candidat: CandidatCreate, current_user: dict = Depends(get_current_user)):
    candidat_obj = Candidat(**candidat.model_dump())
    doc = candidat_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.candidats.insert_one(doc)
    return candidat_obj

@api_router.put("/candidats/{candidat_id}", response_model=Candidat)
async def update_candidat(candidat_id: str, candidat: CandidatUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.candidats.find_one({"id": candidat_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Candidat non trouvé")
    
    update_data = {k: v for k, v in candidat.model_dump().items() if v is not None}
    if update_data:
        await db.candidats.update_one({"id": candidat_id}, {"$set": update_data})
    
    updated = await db.candidats.find_one({"id": candidat_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/candidats/{candidat_id}")
async def delete_candidat(candidat_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.candidats.delete_one({"id": candidat_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Candidat non trouvé")
    # Supprimer aussi les process associés
    await db.process.delete_many({"candidat_id": candidat_id})
    return {"message": "Candidat supprimé"}

# ============ PROCESS ROUTES ============

@api_router.get("/process", response_model=List[dict])
async def get_all_process(current_user: dict = Depends(get_current_user)):
    """Récupère tous les process avec les infos candidat et poste"""
    processes = await db.process.find({}, {"_id": 0}).sort("updated_at", -1).to_list(1000)
    
    result = []
    for proc in processes:
        candidat = await db.candidats.find_one({"id": proc['candidat_id']}, {"_id": 0})
        poste = await db.postes.find_one({"id": proc['poste_id']}, {"_id": 0})
        if candidat and poste:
            result.append({
                **proc,
                "candidat": candidat,
                "poste": poste
            })
    return result

@api_router.get("/process/candidat/{candidat_id}")
async def get_process_by_candidat(candidat_id: str, current_user: dict = Depends(get_current_user)):
    """Récupère tous les process d'un candidat"""
    processes = await db.process.find({"candidat_id": candidat_id}, {"_id": 0}).to_list(100)
    
    result = []
    for proc in processes:
        poste = await db.postes.find_one({"id": proc['poste_id']}, {"_id": 0})
        if poste:
            result.append({**proc, "poste": poste})
    return result

@api_router.get("/process/poste/{poste_id}")
async def get_process_by_poste(poste_id: str, current_user: dict = Depends(get_current_user)):
    """Récupère tous les process d'un poste"""
    processes = await db.process.find({"poste_id": poste_id}, {"_id": 0}).to_list(100)
    
    result = []
    for proc in processes:
        candidat = await db.candidats.find_one({"id": proc['candidat_id']}, {"_id": 0})
        if candidat:
            result.append({**proc, "candidat": candidat})
    return result

@api_router.post("/process", response_model=Process)
async def create_process(process: ProcessCreate, current_user: dict = Depends(get_current_user)):
    # Vérifier que le candidat et le poste existent
    candidat = await db.candidats.find_one({"id": process.candidat_id}, {"_id": 0})
    if not candidat:
        raise HTTPException(status_code=404, detail="Candidat non trouvé")
    
    poste = await db.postes.find_one({"id": process.poste_id}, {"_id": 0})
    if not poste:
        raise HTTPException(status_code=404, detail="Poste non trouvé")
    
    # Vérifier qu'un process n'existe pas déjà pour ce couple
    existing = await db.process.find_one({
        "candidat_id": process.candidat_id,
        "poste_id": process.poste_id
    }, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Un process existe déjà pour ce candidat et ce poste")
    
    process_obj = Process(**process.model_dump())
    doc = process_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.process.insert_one(doc)
    return process_obj

@api_router.put("/process/{process_id}", response_model=dict)
async def update_process(process_id: str, process: ProcessUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.process.find_one({"id": process_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Process non trouvé")
    
    update_data = {k: v for k, v in process.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    if update_data:
        await db.process.update_one({"id": process_id}, {"$set": update_data})
    
    updated = await db.process.find_one({"id": process_id}, {"_id": 0})
    
    # Récupérer les infos candidat et poste
    candidat = await db.candidats.find_one({"id": updated['candidat_id']}, {"_id": 0})
    poste = await db.postes.find_one({"id": updated['poste_id']}, {"_id": 0})
    
    return {**updated, "candidat": candidat, "poste": poste}

@api_router.delete("/process/{process_id}")
async def delete_process(process_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.process.delete_one({"id": process_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Process non trouvé")
    return {"message": "Process supprimé"}

# ============ POSTES ROUTES ============

@api_router.get("/postes", response_model=List[Poste])
async def get_postes(current_user: dict = Depends(get_current_user)):
    postes = await db.postes.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for p in postes:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return postes

@api_router.get("/postes/{poste_id}", response_model=Poste)
async def get_poste(poste_id: str, current_user: dict = Depends(get_current_user)):
    poste = await db.postes.find_one({"id": poste_id}, {"_id": 0})
    if not poste:
        raise HTTPException(status_code=404, detail="Poste non trouvé")
    if isinstance(poste.get('created_at'), str):
        poste['created_at'] = datetime.fromisoformat(poste['created_at'])
    return poste

@api_router.post("/postes", response_model=Poste)
async def create_poste(poste: PosteCreate, current_user: dict = Depends(get_current_user)):
    poste_obj = Poste(**poste.model_dump())
    doc = poste_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.postes.insert_one(doc)
    return poste_obj

@api_router.put("/postes/{poste_id}", response_model=Poste)
async def update_poste(poste_id: str, poste: PosteUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.postes.find_one({"id": poste_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Poste non trouvé")
    
    update_data = {k: v for k, v in poste.model_dump().items() if v is not None}
    if update_data:
        await db.postes.update_one({"id": poste_id}, {"$set": update_data})
    
    updated = await db.postes.find_one({"id": poste_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/postes/{poste_id}")
async def delete_poste(poste_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.postes.delete_one({"id": poste_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Poste non trouvé")
    return {"message": "Poste supprimé"}

# ============ MATCHING ROUTES ============

@api_router.get("/matching/{poste_id}", response_model=List[Match])
async def get_matches_for_poste(poste_id: str, current_user: dict = Depends(get_current_user)):
    poste = await db.postes.find_one({"id": poste_id}, {"_id": 0})
    if not poste:
        raise HTTPException(status_code=404, detail="Poste non trouvé")
    
    candidats = await db.candidats.find({}, {"_id": 0}).to_list(1000)
    matches = []
    
    for candidat in candidats:
        match_result = await calculate_match_score_async(candidat, poste)
        if match_result['score'] > 0:
            if isinstance(candidat.get('created_at'), str):
                candidat['created_at'] = datetime.fromisoformat(candidat['created_at'])
            matches.append(Match(
                candidat=Candidat(**candidat),
                poste_id=poste_id,
                score=match_result['score'],
                titre_match=match_result['titre_match'],
                zone_match=match_result['zone_match']
            ))
    
    # Sort by score descending
    matches.sort(key=lambda x: x.score, reverse=True)
    return matches

@api_router.get("/matching", response_model=List[dict])
async def get_all_matches(current_user: dict = Depends(get_current_user)):
    """Get all matches grouped by poste"""
    postes = await db.postes.find({}, {"_id": 0}).to_list(1000)
    candidats = await db.candidats.find({}, {"_id": 0}).to_list(1000)
    
    all_matches = []
    for poste in postes:
        if isinstance(poste.get('created_at'), str):
            poste['created_at'] = datetime.fromisoformat(poste['created_at'])
        
        poste_matches = []
        for candidat in candidats:
            match_result = await calculate_match_score_async(candidat, poste)
            if match_result['score'] > 0:
                if isinstance(candidat.get('created_at'), str):
                    candidat['created_at'] = datetime.fromisoformat(candidat['created_at'])
                poste_matches.append({
                    'candidat': candidat,
                    'score': match_result['score'],
                    'titre_match': match_result['titre_match'],
                    'zone_match': match_result['zone_match']
                })
        
        poste_matches.sort(key=lambda x: x['score'], reverse=True)
        all_matches.append({
            'poste': poste,
            'matches': poste_matches[:10],  # Top 10 matches per job
            'total_matches': len(poste_matches)
        })
    
    return all_matches

# ============ STATS ROUTES ============

@api_router.get("/stats")
async def get_stats(current_user: dict = Depends(get_current_user)):
    total_candidats = await db.candidats.count_documents({})
    total_postes = await db.postes.count_documents({})
    
    # Calculate total matches
    postes = await db.postes.find({}, {"_id": 0}).to_list(1000)
    candidats = await db.candidats.find({}, {"_id": 0}).to_list(1000)
    
    total_matches = 0
    high_score_matches = 0
    for poste in postes:
        for candidat in candidats:
            match_result = calculate_match_score(candidat, poste)
            if match_result['score'] > 0:
                total_matches += 1
            if match_result['score'] >= 70:
                high_score_matches += 1
    
    # Stats depuis les process
    processes = await db.process.find({}, {"_id": 0}).to_list(1000)
    statuts_count = {}
    total_honoraires = 0
    candidats_places = 0
    
    for proc in processes:
        statut = proc.get('statut', 'ENCV')
        statuts_count[statut] = statuts_count.get(statut, 0) + 1
        if statut == 'PCLT':
            candidats_places += 1
            if proc.get('honoraire'):
                total_honoraires += proc['honoraire']
    
    return {
        "total_candidats": total_candidats,
        "total_postes": total_postes,
        "total_matches": total_matches,
        "high_score_matches": high_score_matches,
        "statuts_count": statuts_count,
        "total_honoraires": total_honoraires,
        "candidats_places": candidats_places,
        "total_process": len(processes)
    }

@api_router.get("/stats/sources")
async def get_sources_stats(current_user: dict = Depends(get_current_user)):
    """Statistiques par source de candidat"""
    candidats = await db.candidats.find({}, {"_id": 0}).to_list(1000)
    processes = await db.process.find({}, {"_id": 0}).to_list(1000)
    
    # Créer un index des process par candidat_id
    process_by_candidat = {}
    for proc in processes:
        cid = proc['candidat_id']
        if cid not in process_by_candidat:
            process_by_candidat[cid] = []
        process_by_candidat[cid].append(proc)
    
    sources_stats = {}
    for source in SOURCES_CANDIDAT:
        sources_stats[source] = {
            "total": 0,
            "places": 0,
            "honoraires": 0
        }
    
    # Ajouter une catégorie "Non renseigné"
    sources_stats["Non renseigné"] = {"total": 0, "places": 0, "honoraires": 0}
    
    for candidat in candidats:
        source = candidat.get('source') or "Non renseigné"
        if source not in sources_stats:
            source = "Non renseigné"
        
        sources_stats[source]["total"] += 1
        
        # Vérifier les process de ce candidat
        candidat_processes = process_by_candidat.get(candidat['id'], [])
        for proc in candidat_processes:
            if proc.get('statut') == 'PCLT':
                sources_stats[source]["places"] += 1
                if proc.get('honoraire'):
                    sources_stats[source]["honoraires"] += proc['honoraire']
    
    # Convertir en liste triée par honoraires
    result = [
        {"source": source, **stats}
        for source, stats in sources_stats.items()
        if stats["total"] > 0
    ]
    result.sort(key=lambda x: x["honoraires"], reverse=True)
    
    return result

@api_router.get("/config/statuts")
async def get_statuts():
    """Retourne la liste des statuts disponibles"""
    return {
        "statuts": [
            {"code": "NOUVEAU", "label": "Nouveau", "color": "gray"},
            {"code": "ENCV", "label": "Envoyé au client", "color": "blue"},
            {"code": "ENTC", "label": "Entretien client", "color": "purple"},
            {"code": "PROPALE", "label": "Sous proposition", "color": "orange"},
            {"code": "PCLT", "label": "Placé", "color": "green"},
            {"code": "REFUS", "label": "Refus propale", "color": "red"},
            {"code": "NOGO_DISPO", "label": "Plus disponible", "color": "gray"}
        ],
        "sources": SOURCES_CANDIDAT
    }

# ============ EXPORT EXCEL ROUTES ============

@api_router.get("/export/candidats")
async def export_candidats_excel(current_user: dict = Depends(get_current_user)):
    """Export des candidats en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    candidats = await db.candidats.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    processes = await db.process.find({}, {"_id": 0}).to_list(10000)
    
    # Index des process par candidat
    process_by_candidat = {}
    for proc in processes:
        cid = proc['candidat_id']
        if cid not in process_by_candidat:
            process_by_candidat[cid] = []
        process_by_candidat[cid].append(proc)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidats"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Prénom", "Nom", "Poste recherché", "Ville", "Rayon (km)", 
               "Rémunération", "Disponibilité", "Source", "Process actifs", "Date création"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row, candidat in enumerate(candidats, 2):
        procs = process_by_candidat.get(candidat['id'], [])
        active_procs = [p for p in procs if p['statut'] not in ['PCLT', 'REFUS', 'NOGO_DISPO']]
        
        data = [
            candidat.get('prenom', ''),
            candidat.get('nom', ''),
            candidat.get('titre_poste', ''),
            candidat.get('ville', ''),
            candidat.get('rayon_km', 30),
            candidat.get('remuneration', ''),
            candidat.get('disponibilite', ''),
            candidat.get('source', ''),
            len(active_procs),
            candidat.get('created_at', '')[:10] if candidat.get('created_at') else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
    
    # Adjust column widths
    column_widths = [12, 15, 25, 15, 12, 15, 15, 20, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"candidats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/postes")
async def export_postes_excel(current_user: dict = Depends(get_current_user)):
    """Export des postes en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    postes = await db.postes.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    processes = await db.process.find({}, {"_id": 0}).to_list(10000)
    
    # Index des process par poste
    process_by_poste = {}
    for proc in processes:
        pid = proc['poste_id']
        if pid not in process_by_poste:
            process_by_poste[pid] = []
        process_by_poste[pid].append(proc)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Postes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    orange_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Entreprise", "Poste", "Ville", "Contact", "Email", "Convention", "Process en cours", 
               "Candidats placés", "Date création"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row, poste in enumerate(postes, 2):
        procs = process_by_poste.get(poste['id'], [])
        active_procs = [p for p in procs if p['statut'] not in ['PCLT', 'REFUS', 'NOGO_DISPO']]
        placed = [p for p in procs if p['statut'] == 'PCLT']
        
        data = [
            poste.get('entreprise', ''),
            poste.get('titre_poste', ''),
            poste.get('ville', ''),
            poste.get('contact', ''),
            poste.get('email_contact', ''),
            "Oui" if poste.get('convention_signee') else "Non",
            len(active_procs),
            len(placed),
            poste.get('created_at', '')[:10] if poste.get('created_at') else ''
        ]
        
        row_fill = green_fill if poste.get('convention_signee') else orange_fill
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
            cell.fill = row_fill
    
    # Adjust column widths
    column_widths = [20, 25, 15, 20, 25, 12, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"postes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/process")
async def export_process_excel(current_user: dict = Depends(get_current_user)):
    """Export des process en Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    processes = await db.process.find({}, {"_id": 0}).sort("updated_at", -1).to_list(10000)
    candidats = {c['id']: c for c in await db.candidats.find({}, {"_id": 0}).to_list(10000)}
    postes = {p['id']: p for p in await db.postes.find({}, {"_id": 0}).to_list(10000)}
    
    STATUT_LABELS = {
        "ENCV": "Envoyé au client",
        "ENTC": "Entretien client", 
        "PROPALE": "Sous proposition",
        "PCLT": "Placé",
        "REFUS": "Refus",
        "NOGO_DISPO": "Plus disponible"
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Process"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Candidat", "Poste candidat", "Entreprise", "Poste entreprise", 
               "Ville", "Statut", "Honoraire (€)", "Notes", "Dernière MAJ"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row, proc in enumerate(processes, 2):
        candidat = candidats.get(proc['candidat_id'], {})
        poste = postes.get(proc['poste_id'], {})
        
        data = [
            f"{candidat.get('prenom', '')} {candidat.get('nom', '')}",
            candidat.get('titre_poste', ''),
            poste.get('entreprise', ''),
            poste.get('titre_poste', ''),
            poste.get('ville', ''),
            STATUT_LABELS.get(proc.get('statut', ''), proc.get('statut', '')),
            proc.get('honoraire', ''),
            proc.get('notes', ''),
            proc.get('updated_at', '')[:10] if proc.get('updated_at') else ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left')
    
    # Adjust column widths
    column_widths = [20, 20, 20, 20, 15, 18, 12, 30, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"process_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ IMPORT EXCEL ROUTES ============

@api_router.post("/import/candidats")
async def import_candidats_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Import des candidats depuis Excel"""
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format Excel (.xlsx)")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        # Lire les headers
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        
        # Mapping des colonnes
        col_map = {}
        for i, h in enumerate(headers):
            if 'prénom' in h or 'prenom' in h:
                col_map['prenom'] = i
            elif 'candidat' in h:
                col_map['candidat'] = i  # Nom complet
            elif 'nom' in h and 'prénom' not in h and 'prenom' not in h:
                col_map['nom'] = i
            elif 'poste' in h and 'recherch' in h:
                col_map['titre_poste'] = i
            elif 'poste' in h or 'titre' in h:
                col_map['titre_poste'] = col_map.get('titre_poste', i)
            elif 'ville' in h:
                col_map['ville'] = i
            elif 'département' in h or 'departement' in h or 'dept' in h:
                col_map['departement'] = i
            elif 'code' in h and 'postal' in h or h == 'cp' or h == 'code postal':
                col_map['code_postal'] = i
            elif 'rayon' in h or 'km' in h:
                col_map['rayon_km'] = i
            elif 'rémunération' in h or 'remuneration' in h or 'salaire' in h or h == 'rem':
                col_map['remuneration'] = i
            elif 'dispo' in h or 'disponibilité' in h or 'disponibilite' in h:
                col_map['disponibilite'] = i
            elif 'source' in h:
                col_map['source'] = i
            elif 'réf' in h or 'ref' in h:
                col_map['ref'] = i
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extraire prénom et nom
                prenom = ''
                nom = ''
                
                # Si on a une colonne "Candidats" avec nom complet
                if col_map.get('candidat') is not None and row[col_map.get('candidat')]:
                    nom_complet = str(row[col_map.get('candidat')]).strip()
                    parts = nom_complet.split(' ', 1)
                    if len(parts) >= 2:
                        prenom = parts[0]
                        nom = parts[1]
                    else:
                        nom = nom_complet
                        prenom = '-'
                else:
                    # Sinon colonnes séparées
                    prenom = str(row[col_map.get('prenom', 0)] or '').strip() if col_map.get('prenom') is not None and len(row) > col_map.get('prenom', 0) and row[col_map.get('prenom')] else ''
                    nom = str(row[col_map.get('nom', 1)] or '').strip() if col_map.get('nom') is not None and len(row) > col_map.get('nom', 1) and row[col_map.get('nom')] else ''
                
                if not prenom and not nom:
                    continue  # Ligne vide
                
                if not prenom:
                    prenom = '-'
                if not nom:
                    nom = '-'
                
                titre_poste = str(row[col_map.get('titre_poste')] or '').strip() if col_map.get('titre_poste') is not None and len(row) > col_map.get('titre_poste') else ''
                ville = str(row[col_map.get('ville')] or '').strip() if col_map.get('ville') is not None and len(row) > col_map.get('ville') else ''
                
                # Si pas de ville mais un département
                if not ville and col_map.get('departement') is not None and len(row) > col_map.get('departement'):
                    ville = str(row[col_map.get('departement')] or '').strip()
                
                code_postal = str(row[col_map.get('code_postal')] or '').strip() if col_map.get('code_postal') is not None and len(row) > col_map.get('code_postal') else ''
                
                rayon_km = 30
                if col_map.get('rayon_km') is not None and len(row) > col_map.get('rayon_km') and row[col_map.get('rayon_km')]:
                    try:
                        val = str(row[col_map.get('rayon_km')]).replace('km', '').replace('KM', '').strip()
                        rayon_km = int(float(val))
                    except:
                        pass
                
                remuneration = str(row[col_map.get('remuneration')] or '').strip() if col_map.get('remuneration') is not None and len(row) > col_map.get('remuneration') else ''
                disponibilite = str(row[col_map.get('disponibilite')] or '').strip() if col_map.get('disponibilite') is not None and len(row) > col_map.get('disponibilite') else ''
                source = str(row[col_map.get('source')] or '').strip() if col_map.get('source') is not None and len(row) > col_map.get('source') else None
                
                # Créer le candidat
                candidat_doc = {
                    "id": str(uuid.uuid4()),
                    "prenom": prenom,
                    "nom": nom,
                    "titre_poste": titre_poste or "Non renseigné",
                    "ville": ville or "Non renseigné",
                    "code_postal": code_postal or None,
                    "rayon_km": rayon_km,
                    "remuneration": remuneration or None,
                    "disponibilite": disponibilite or None,
                    "source": source,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.candidats.insert_one(candidat_doc)
                imported += 1
                
            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "errors": errors[:10]  # Max 10 erreurs
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier: {str(e)}")

@api_router.post("/import/postes")
async def import_postes_excel(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Import des postes depuis Excel"""
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format Excel (.xlsx)")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        # Lire les headers
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        
        # Mapping des colonnes
        col_map = {}
        for i, h in enumerate(headers):
            if 'entreprise' in h or 'société' in h or 'societe' in h or 'client' in h:
                col_map['entreprise'] = i
            elif 'poste' in h or 'titre' in h or 'intitulé' in h:
                col_map['titre_poste'] = col_map.get('titre_poste', i)
            elif 'ville' in h or 'lieu' in h or 'localisation' in h:
                col_map['ville'] = i
            elif 'code' in h and 'postal' in h or h == 'cp' or h == 'code postal':
                col_map['code_postal'] = i
            elif 'convention' in h:
                col_map['convention_signee'] = i
            elif 'contact' in h and 'email' not in h and 'mail' not in h:
                col_map['contact'] = i
            elif 'email' in h or 'mail' in h:
                col_map['email_contact'] = i
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                entreprise = str(row[col_map.get('entreprise', 0)] or '').strip() if col_map.get('entreprise') is not None and row[col_map.get('entreprise')] else ''
                
                if not entreprise:
                    continue  # Ligne vide
                
                titre_poste = str(row[col_map.get('titre_poste', 1)] or '').strip() if col_map.get('titre_poste') is not None else ''
                ville = str(row[col_map.get('ville', 2)] or '').strip() if col_map.get('ville') is not None else ''
                
                convention_signee = False
                if col_map.get('convention_signee') is not None and row[col_map.get('convention_signee')]:
                    val = str(row[col_map.get('convention_signee')]).lower().strip()
                    convention_signee = val in ('oui', 'yes', 'true', '1', 'signée', 'signee')
                
                contact = str(row[col_map.get('contact')] or '').strip() if col_map.get('contact') is not None else ''
                email_contact = str(row[col_map.get('email_contact')] or '').strip() if col_map.get('email_contact') is not None else ''
                code_postal = str(row[col_map.get('code_postal')] or '').strip() if col_map.get('code_postal') is not None else ''
                
                # Créer le poste
                poste_doc = {
                    "id": str(uuid.uuid4()),
                    "entreprise": entreprise,
                    "titre_poste": titre_poste or "Non renseigné",
                    "ville": ville or "Non renseigné",
                    "code_postal": code_postal or None,
                    "convention_signee": convention_signee,
                    "contact": contact or None,
                    "email_contact": email_contact or None,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.postes.insert_one(poste_doc)
                imported += 1
                
            except Exception as e:
                errors.append(f"Ligne {row_idx}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "errors": errors[:10]
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de la lecture du fichier: {str(e)}")

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
